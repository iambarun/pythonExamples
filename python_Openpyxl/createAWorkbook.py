import openpyxl as xl
wb = xl.Workbook()
ws = wb.active
ws = wb.create_sheet("FirstSheet", 0)
ws.sheet_properties.tabColor = '1072BA'
d = ws.cell(row=4, column=2, value=12)
c = ws['B4']
print(c.value)
ws1 = wb.create_sheet("ZeroIndex", 0)
wb.save('files/test.xlsx')